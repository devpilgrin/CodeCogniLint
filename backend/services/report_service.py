from io import BytesIO
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CATEGORY_LABEL = {
    "syntax": "Синтаксис",
    "semantic": "Семантика",
    "analysis": "Анализ",
}
SEVERITY_LABEL = {
    "critical": "Критично",
    "warning": "Важно",
    "info": "Инфо",
}
SEVERITY_COLOR = {
    "critical": "FFD9534F",
    "warning":  "FFF0AD4E",
    "info":     "FF5BC0DE",
}


def _aggregate(results: dict) -> dict:
    """Compute summary statistics from {path: AnalysisResult}."""
    total_files = len(results)
    by_category: dict[str, int] = {}
    by_severity: dict[str, int] = {}
    per_file = []
    total_violations = 0

    for path, res in results.items():
        if not isinstance(res, dict):
            continue  # фаззер/битые данные: пропускаем некорректные записи
        violations = (res or {}).get("violations") or []
        violations = [v for v in violations if isinstance(v, dict)]
        per_file.append({
            "path": path,
            "total": len(violations),
            "critical": sum(1 for v in violations if v.get("severity") == "critical"),
            "warning":  sum(1 for v in violations if v.get("severity") == "warning"),
            "info":     sum(1 for v in violations if v.get("severity") == "info"),
            "violations": violations,
        })
        total_violations += len(violations)
        for v in violations:
            cat = v.get("category", "—")
            sev = v.get("severity", "—")
            by_category[cat] = by_category.get(cat, 0) + 1
            by_severity[sev] = by_severity.get(sev, 0) + 1

    per_file.sort(key=lambda x: x["total"], reverse=True)
    return {
        "total_files": total_files,
        "total_violations": total_violations,
        "by_category": by_category,
        "by_severity": by_severity,
        "per_file": per_file,
    }


def _xsafe(v) -> str:
    """Безопасное значение для ячейки XLSX: openpyxl падает на недопустимых
    XML-символах (управляющие/суррогаты из фаззера) — вычищаем."""
    s = v if isinstance(v, str) else str(v)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff]", "?", s)


def generate_xlsx(results: dict) -> bytes:
    agg = _aggregate(results)
    wb = Workbook()

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FF22272E")
    header_font = Font(bold=True, color="FFFFFFFF")
    thin = Side(border_style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14, color="FF1F6FEB")

    # --- Sheet 1: Сводка ---
    ws = wb.active
    ws.title = "Сводка"

    ws["A1"] = "Сводный отчёт по сканированию"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="FF888888")
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="Параметр").font = bold
    ws.cell(row=row, column=2, value="Значение").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Файлов проанализировано")
    ws.cell(row=row, column=2, value=agg["total_files"])
    row += 1
    ws.cell(row=row, column=1, value="Всего нарушений")
    ws.cell(row=row, column=2, value=agg["total_violations"])

    row += 2
    ws.cell(row=row, column=1, value="По категории").font = bold
    row += 1
    for cat, count in sorted(agg["by_category"].items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=_xsafe(CATEGORY_LABEL.get(cat, cat)))
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="По серьёзности").font = bold
    row += 1
    for sev, count in sorted(agg["by_severity"].items(), key=lambda x: -x[1]):
        cell = ws.cell(row=row, column=1, value=_xsafe(SEVERITY_LABEL.get(sev, sev)))
        if sev in SEVERITY_COLOR:
            cell.font = Font(bold=True, color=SEVERITY_COLOR[sev].replace("FF", "FF", 1))
        ws.cell(row=row, column=2, value=count)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14

    # --- Sheet 2: По файлам ---
    ws2 = wb.create_sheet("По файлам")
    headers = ["Файл", "Всего", "Критично", "Важно", "Инфо"]
    for col_idx, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for r_idx, f in enumerate(agg["per_file"], start=2):
        ws2.cell(row=r_idx, column=1, value=_xsafe(f["path"])).border = border
        ws2.cell(row=r_idx, column=2, value=f["total"]).border = border
        crit = ws2.cell(row=r_idx, column=3, value=f["critical"])
        if f["critical"] > 0: crit.font = Font(bold=True, color=SEVERITY_COLOR["critical"])
        crit.border = border
        warn = ws2.cell(row=r_idx, column=4, value=f["warning"])
        if f["warning"] > 0: warn.font = Font(color=SEVERITY_COLOR["warning"])
        warn.border = border
        info = ws2.cell(row=r_idx, column=5, value=f["info"])
        if f["info"] > 0: info.font = Font(color=SEVERITY_COLOR["info"])
        info.border = border
        for col in range(2, 6):
            ws2.cell(row=r_idx, column=col).alignment = Alignment(horizontal="center")

    for col_letter, w in zip("ABCDE", [50, 10, 12, 10, 10]):
        ws2.column_dimensions[col_letter].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Все нарушения ---
    ws3 = wb.create_sheet("Нарушения")
    headers = ["Файл", "Строка", "Severity", "Категория", "Правило", "Описание", "Сниппет", "Рекомендация"]
    for col_idx, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    r_idx = 2
    for f in agg["per_file"]:
        for v in f["violations"]:
            ls = v.get("line_start", "")
            le = v.get("line_end", "")
            line_str = f"{ls}" if not le or le == ls else f"{ls}–{le}"

            ws3.cell(row=r_idx, column=1, value=_xsafe(f["path"]))
            ws3.cell(row=r_idx, column=2, value=_xsafe(line_str))
            sev_cell = ws3.cell(row=r_idx, column=3, value=_xsafe(SEVERITY_LABEL.get(v.get("severity", ""), v.get("severity", ""))))
            sev_key = v.get("severity")
            if sev_key in SEVERITY_COLOR:
                sev_cell.font = Font(bold=True, color=SEVERITY_COLOR[sev_key])
            ws3.cell(row=r_idx, column=4, value=_xsafe(CATEGORY_LABEL.get(v.get("category", ""), v.get("category", ""))))
            ws3.cell(row=r_idx, column=5, value=_xsafe(v.get("rule_description", "")))
            ws3.cell(row=r_idx, column=6, value=_xsafe(v.get("explanation", "")))
            ws3.cell(row=r_idx, column=7, value=_xsafe(v.get("code_snippet", "")))
            ws3.cell(row=r_idx, column=8, value=_xsafe(v.get("suggestion", "")))

            for col in range(1, 9):
                cell = ws3.cell(row=r_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            r_idx += 1

    for col_letter, w in zip("ABCDEFGH", [32, 10, 12, 14, 40, 50, 40, 50]):
        ws3.column_dimensions[col_letter].width = w
    ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_md(results: dict) -> str:
    agg = _aggregate(results)
    out: list[str] = []
    out.append("# Сводный отчёт по сканированию")
    out.append("")
    out.append(f"_Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}_")
    out.append("")

    out.append("## Сводка")
    out.append("")
    out.append(f"- **Файлов проанализировано:** {agg['total_files']}")
    out.append(f"- **Всего нарушений:** {agg['total_violations']}")
    out.append("")

    if agg["by_category"]:
        out.append("### По категории")
        for cat, count in sorted(agg["by_category"].items(), key=lambda x: -x[1]):
            out.append(f"- {CATEGORY_LABEL.get(cat, cat)}: **{count}**")
        out.append("")

    if agg["by_severity"]:
        out.append("### По серьёзности")
        for sev, count in sorted(agg["by_severity"].items(), key=lambda x: -x[1]):
            out.append(f"- {SEVERITY_LABEL.get(sev, sev)}: **{count}**")
        out.append("")

    out.append("## Сводка по файлам")
    out.append("")
    out.append("| Файл | Всего | Критично | Важно | Инфо |")
    out.append("|------|-------|----------|-------|------|")
    for f in agg["per_file"]:
        out.append(f"| `{f['path']}` | {f['total']} | {f['critical']} | {f['warning']} | {f['info']} |")
    out.append("")

    files_with_violations = [f for f in agg["per_file"] if f["total"] > 0]
    if files_with_violations:
        out.append("## Детализация нарушений")
        out.append("")
        for f in files_with_violations:
            out.append(f"### `{f['path']}` — {f['total']} нарушений")
            out.append("")
            for v in f["violations"]:
                ls = v.get("line_start", "?")
                le = v.get("line_end", ls)
                line_str = f"{ls}" if le == ls else f"{ls}–{le}"
                sev = SEVERITY_LABEL.get(v.get("severity", ""), v.get("severity", ""))
                cat = CATEGORY_LABEL.get(v.get("category", ""), v.get("category", ""))
                out.append(f"#### Строка {line_str} — {sev} / {cat}")
                out.append("")
                if v.get("rule_description"):
                    out.append(f"**{v['rule_description']}**")
                    out.append("")
                if v.get("code_snippet"):
                    out.append("```")
                    out.append(v["code_snippet"])
                    out.append("```")
                    out.append("")
                if v.get("explanation"):
                    out.append(v["explanation"])
                    out.append("")
                if v.get("suggestion"):
                    out.append(f"> 💡 **Рекомендация:** {v['suggestion']}")
                    out.append("")
            out.append("---")
            out.append("")

    return "\n".join(out)


# ---- HTML-экспорт отчёта аудита (волна: открытые направления) ----

_AUDIT_RISK_LABEL = {
    "low": ("Низкий", "#2ea043"),
    "medium": ("Средний", "#d29922"),
    "high": ("Высокий", "#f85149"),
    "critical": ("Критический", "#da3633"),
    "none": ("Чисто", "#2ea043"),
}


def _esc(v) -> str:
    s = v if isinstance(v, str) else str(v)
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;"))


def generate_audit_html(report: dict) -> str:
    """Статичный HTML-отчёт аудита из готового JSON (детерминированный рендер)."""
    syn = report.get("synthesis") or {}
    risk = (syn or {}).get("risk_level") or ("none" if not report.get("total_findings") else "medium")
    risk_label, risk_color = _AUDIT_RISK_LABEL.get(risk, _AUDIT_RISK_LABEL["medium"])
    tools = ", ".join(k for k, v in (report.get("tools") or {}).items() if v) or "—"

    def li(items):
        return "\n".join(f"<li>{_esc(i)}</li>" for i in (items or [])) or "<li>—</li>"

    matrix_rows = "\n".join(
        f"<tr><td>{_esc(m['cwe'])}</td><td>{_esc(m['severity'])}</td>"
        f"<td>{m['count']}</td><td>{_esc(m['exploitability'])}</td></tr>"
        for m in (report.get("matrix") or [])
    ) or "<tr><td colspan='4'>нет данных</td></tr>"

    domains_html = []
    for d in (report.get("domains") or []):
        a = d.get("agent") or {}
        _, dc = _AUDIT_RISK_LABEL.get(a.get("risk_level", "medium"), _AUDIT_RISK_LABEL["medium"])
        fps = a.get("false_positives") or []
        domains_html.append(f"""
<section class="card">
  <h3>{_esc(d.get('label', d.get('domain', '?')))}
      <span class="badge" style="background:{dc}">{_esc(a.get('risk_level', '—'))}</span>
      <small>{d.get('findings_count', 0)} находок</small></h3>
  <p>{_esc(a.get('assessment', ''))}</p>
  <h4>Эксплуатируемость</h4><ul>{li(a.get('exploitability'))}</ul>
  <h4>Рекомендации</h4><ul>{li(a.get('recommendations'))}</ul>
  {f"<h4>Ложные срабатывания</h4><ul>{li(fps)}</ul>" if fps else ""}
</section>""")

    return f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<title>Аудит безопасности — {_esc(report.get('workspace', ''))}</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; background: #0d1117; color: #c9d1d9;
       max-width: 960px; margin: 24px auto; padding: 0 16px; }}
h1 {{ font-size: 22px; }} h3 {{ margin-bottom: 6px; }} h4 {{ margin: 10px 0 4px; font-size: 13px; color: #8b949e; }}
.card {{ background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 14px 18px; margin: 12px 0; }}
.badge {{ color: #fff; border-radius: 10px; padding: 2px 10px; font-size: 12px; vertical-align: middle; }}
table {{ border-collapse: collapse; width: 100%; }} td, th {{ border: 1px solid #30363d; padding: 6px 10px; font-size: 13px; }}
th {{ background: #21262d; text-align: left; }} small {{ color: #8b949e; font-weight: normal; }}
ul {{ margin: 4px 0; padding-left: 20px; }} li {{ margin: 3px 0; font-size: 13px; }}
.meta {{ color: #8b949e; font-size: 12px; }}
</style></head><body>
<h1>Отчёт аудита безопасности <span class="badge" style="background:{risk_color}">{risk_label}</span></h1>
<p class="meta">Workspace: {_esc(report.get('workspace', ''))} · Сформирован: {_esc(report.get('scanned_at', ''))} ·
Инструменты: {_esc(tools)} · Находок: {report.get('total_findings', 0)}
(+{report.get('new_findings', 0)} новых / −{report.get('fixed_findings', 0)} исправлено)</p>

{f'''<section class="card"><h3>Итоговый вердикт</h3><p>{_esc(syn.get('verdict', ''))}</p>
<h4>Векторы атаки</h4><ul>{li(syn.get('attack_vectors'))}</ul>
<h4>Приоритетные действия</h4><ul>{li(syn.get('top_actions'))}</ul></section>''' if syn else ""}

<section class="card"><h3>Матрица рисков</h3>
<table><tr><th>CWE</th><th>Тяжесть</th><th>Кол-во</th><th>Эксплуатируемость</th></tr>{matrix_rows}</table>
</section>

{''.join(domains_html) or '<section class="card"><p>Находок нет — аудит чистый.</p></section>'}
<p class="meta">Сформировано CodeCogniLint · детерминированный рендер из JSON отчёта</p>
</body></html>"""

